"""Import the curated hackathon XLSX (embedded before/after images + metadata).

The spreadsheet stores, per case, a Before image (column 0) and an After image
(column 1) anchored on the same row, plus Case description, Age, and Biologic
columns. Every row is a *success* case (the patient improved on the listed
biologic). This script extracts the images to ``data/images/<CASE_ID>/`` and
writes the 4-column ``data/cases.csv`` consumed by the matching engine.

Usage:
    python scripts/import_dataset.py "/path/to/Dataset HACKATHON '26 (5).xlsx"
"""

from __future__ import annotations

import csv
import io
import re
import shutil
import sys
import warnings
from collections import defaultdict
from pathlib import Path

import openpyxl
from PIL import Image

warnings.filterwarnings("ignore")

DATA_ROOT = Path(__file__).resolve().parents[1] / "data"
IMAGES_ROOT = DATA_ROOT / "images"
CSV_PATH = DATA_ROOT / "cases.csv"

# Pipeline preprocessing minimums (see app/services/preprocessing.py).
MIN_SHORT_EDGE = 256
MIN_LONG_EDGE = 512

BIOLOGIC_COL = 5  # 1-indexed spreadsheet columns
AGE_COL = 4
DESC_COL = 3


def parse_age(age_cell: object, desc: str) -> str:
    """Return an integer age as a string, or '' when unknown."""
    for source in (age_cell, desc):
        if not source:
            continue
        match = re.search(r"(\d{1,3})\s*year", str(source))
        if match:
            value = int(match.group(1))
            if 0 < value < 120:
                return str(value)
    return ""


def to_rgb(data: bytes) -> Image.Image:
    img = Image.open(io.BytesIO(data))
    if img.mode in ("RGBA", "LA", "P"):
        img = img.convert("RGBA")
        background = Image.new("RGBA", img.size, (255, 255, 255, 255))
        img = Image.alpha_composite(background, img).convert("RGB")
    else:
        img = img.convert("RGB")
    return img


def upscale_to_minimums(img: Image.Image) -> Image.Image:
    w, h = img.size
    short, long = min(w, h), max(w, h)
    scale = max(1.0, MIN_SHORT_EDGE / short, MIN_LONG_EDGE / long)
    if scale > 1.0:
        img = img.resize((round(w * scale), round(h * scale)), Image.LANCZOS)
    return img


def main() -> None:
    if len(sys.argv) < 2:
        raise SystemExit("Usage: python scripts/import_dataset.py <xlsx path>")
    xlsx_path = Path(sys.argv[1])
    if not xlsx_path.exists():
        raise SystemExit(f"File not found: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.worksheets[0]

    # Group embedded images by anchor row: {row: {col: bytes}}.
    by_row: dict[int, dict[int, bytes]] = defaultdict(dict)
    for image in ws._images:  # noqa: SLF001 - openpyxl exposes images here
        anchor = image.anchor._from  # noqa: SLF001
        raw = image._data() if callable(image._data) else image.ref  # noqa: SLF001
        by_row[anchor.row][anchor.col] = raw

    if IMAGES_ROOT.exists():
        shutil.rmtree(IMAGES_ROOT)
    IMAGES_ROOT.mkdir(parents=True, exist_ok=True)

    rows_out: list[dict[str, str]] = []
    case_index = 0
    for anchor_row in sorted(by_row):
        cells = by_row[anchor_row]
        if 0 not in cells or 1 not in cells:
            print(f"skip anchor row {anchor_row}: missing before/after image")
            continue
        sheet_row = anchor_row + 1  # anchor rows are 0-indexed
        biologic = ws.cell(sheet_row, BIOLOGIC_COL).value
        desc = ws.cell(sheet_row, DESC_COL).value or ""
        age_cell = ws.cell(sheet_row, AGE_COL).value
        if not biologic or str(biologic).strip().lower() == "biologic":
            print(f"skip sheet row {sheet_row}: no biologic value")
            continue
        biologic = str(biologic).strip().title()

        case_index += 1
        case_id = f"DM-{case_index:03d}"
        case_dir = IMAGES_ROOT / case_id
        case_dir.mkdir(parents=True, exist_ok=True)

        before = upscale_to_minimums(to_rgb(cells[0]))
        after = upscale_to_minimums(to_rgb(cells[1]))
        before.save(case_dir / "before.jpg", "JPEG", quality=92)
        after.save(case_dir / "after.jpg", "JPEG", quality=92)

        rows_out.append(
            {
                "Before": f"images/{case_id}/before.jpg",
                "After": f"images/{case_id}/after.jpg",
                "Biologic": biologic,
                "Age": parse_age(age_cell, desc),
            }
        )
        print(f"{case_id}: {biologic} age={rows_out[-1]['Age'] or '?'}")

    with CSV_PATH.open("w", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["Before", "After", "Biologic", "Age"])
        writer.writeheader()
        writer.writerows(rows_out)

    biologics = defaultdict(int)
    for row in rows_out:
        biologics[row["Biologic"]] += 1
    print(f"\nWrote {len(rows_out)} cases to {CSV_PATH}")
    print("Biologic distribution:", dict(biologics))


if __name__ == "__main__":
    main()
